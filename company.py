import openpyxl
import requests
import time
import logging
import random
import json
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import os

# Set up logging
logging.basicConfig(filename='company_info_extraction.log', level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

# Load environment variables
load_dotenv()

# ProxyCurl API endpoints and your API key
LOOKUP_ENDPOINT = "https://nubela.co/proxycurl/api/linkedin/company/resolve"
PROFILE_ENDPOINT = "https://nubela.co/proxycurl/api/linkedin/company"
PROXYCURL_API = os.getenv("PROXYCURL_API")
logging.info(f"API Key: {PROXYCURL_API[:5]}...")  # Log first 5 characters of API key for verification

# Rate limiting variables
RATE_LIMIT = 2  # requests per minute
RATE_LIMIT_PERIOD = 60  # seconds
last_request_time = 0

def wait_for_rate_limit():
    global last_request_time
    current_time = time.time()
    time_since_last_request = current_time - last_request_time
    if time_since_last_request < RATE_LIMIT_PERIOD / RATE_LIMIT:
        sleep_time = (RATE_LIMIT_PERIOD / RATE_LIMIT) - time_since_last_request
        logging.info(f"Waiting for {sleep_time:.2f} seconds to respect rate limit")
        time.sleep(sleep_time)
    last_request_time = time.time()

def exponential_backoff(attempt, max_delay=300):
    delay = min(60 * (2 ** attempt) + random.uniform(0, 1), max_delay)
    logging.info(f"Rate limit reached. Backing off for {delay:.2f} seconds")
    time.sleep(delay)

def make_api_request(url, headers, params, max_retries=3):
    for attempt in range(max_retries):
        wait_for_rate_limit()
        try:
            response = requests.get(url, headers=headers, params=params)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.HTTPError as e:
            if response.status_code == 429:
                logging.warning(f"Rate limit reached. Attempt {attempt + 1} of {max_retries}")
                exponential_backoff(attempt)
            else:
                logging.error(f"HTTP error occurred: {str(e)}")
                return None
        except requests.exceptions.RequestException as e:
            logging.error(f"An error occurred: {str(e)}")
            return None
    logging.error("Max retries reached. Skipping this request.")
    return None

def lookup_company_url(company_name):
    headers = {'Authorization': f'Bearer {PROXYCURL_API}'}
    params = {'company_name': company_name, 'enrich_profile': 'false'}
    
    result = make_api_request(LOOKUP_ENDPOINT, headers, params)
    if result:
        return result.get('url')
    return None

def get_company_info(company_url):
    headers = {'Authorization': f'Bearer {PROXYCURL_API}'}
    params = {
        'url': company_url,
        'categories': 'include',
        'funding_data': 'include',
        'exit_data': 'include',
        'acquisitions': 'include',
        'extra': 'include',
        'use_cache': 'if-present',
        'fallback_to_cache': 'on-error'
    }
    
    return make_api_request(PROFILE_ENDPOINT, headers, params)

def flatten_dict(d, parent_key='', sep='_'):
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, list):
            items.append((new_key, json.dumps(v)))
        else:
            items.append((new_key, str(v)))
    return dict(items)

# Read company names from the input Excel file
input_workbook = openpyxl.load_workbook('busa3021.xlsx')
input_sheet = input_workbook['Sheet2']

company_names = [cell.value for cell in input_sheet['A'][1:] if cell.value]  # Assuming company names are in column A

# Create a new workbook for output
output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active
output_sheet.title = "Company Information"

# Initialize headers
headers = ["Company Name", "LinkedIn URL", "Status", "Error Details"]
header_row = 1
for col, header in enumerate(headers, start=1):
    output_sheet.cell(row=header_row, column=col, value=header)

# Process companies and update Excel file in real-time
for row, company_name in enumerate(company_names, start=2):
    logging.info(f"Processing: {company_name}")
    
    output_sheet.cell(row=row, column=1, value=company_name)
    
    company_url = lookup_company_url(company_name)
    if not company_url:
        logging.error(f"Could not find LinkedIn URL for {company_name}")
        output_sheet.cell(row=row, column=3, value="URL not found")
        output_sheet.cell(row=row, column=4, value="Company LinkedIn profile not found")
        output_workbook.save('company_information_full.xlsx')
        continue
    
    output_sheet.cell(row=row, column=2, value=company_url)
    
    company_info = get_company_info(company_url)
    
    if company_info:
        flattened_info = flatten_dict(company_info)
        output_sheet.cell(row=row, column=3, value="Data fetched successfully")
        
        # Dynamically add new columns for each piece of data
        for key, value in flattened_info.items():
            if key not in headers:
                headers.append(key)
                col = len(headers)
                output_sheet.cell(row=header_row, column=col, value=key)
            else:
                col = headers.index(key) + 1
            
            output_sheet.cell(row=row, column=col, value=value)
    else:
        logging.error(f"Could not fetch information for {company_name}")
        output_sheet.cell(row=row, column=3, value="Data fetch failed")
        output_sheet.cell(row=row, column=4, value="API request failed or returned no data")
    
    # Save after each company
    output_workbook.save('company_information_full.xlsx')
    logging.info(f"Updated information for {company_name}")

logging.info("Process completed. Final results saved in 'company_information_full.xlsx'")
print("Process completed. Check 'company_information_full.xlsx' for results and 'company_info_extraction.log' for details.")